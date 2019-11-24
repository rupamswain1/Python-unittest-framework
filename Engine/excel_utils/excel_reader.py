import openpyxl
import os
from Engine.utilities.dir_path_provider import dir_path
class read_xlsx:


    def read_locator_excel(self,sheet_name):
        locator_repo_path = "/Resources/locators/locator_repo.xlsx"
        wb_obj=openpyxl.load_workbook(dir_path().get_path()+locator_repo_path)
        sheet_obj=wb_obj[sheet_name]
        max_rows=sheet_obj.max_row
        max_colum=sheet_obj.max_column
        sheet_data=[]
        row_data={}
        for row_num in range(2,max_rows+1):
            for col_num in range(1,max_colum+1):
                row_data.__setitem__(sheet_obj.cell(row=1,column=col_num).value,sheet_obj.cell(row=row_num,column=col_num).value)
            sheet_data.append(row_data)
        return sheet_data